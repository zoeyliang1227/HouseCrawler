import time

import bs4
import hyperlink
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

timeout = 20


def get_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')                 # 瀏覽器不提供可視化頁面
    options.add_argument('-no-sandbox')               # 以最高權限運行
    options.add_argument('--start-maximized')        # 縮放縮放（全屏窗口）設置元素比較準確
    options.add_argument('--disable-gpu')            # 谷歌文檔說明需要加上這個屬性來規避bug
    options.add_argument('--window-size=1920,1080')  # 設置瀏覽器按鈕（窗口大小）
    options.add_argument('--incognito')               # 啟動無痕

    driver = webdriver.Chrome(chrome_options=options)
    driver.get(
        'https://rent.591.com.tw/?rentprice=8000,40000&option=cold,washer,icebox,hotwater,naturalgas,broadband,bed&showMore=1&area=10,&multiRoom=2,3')

    return driver


def search():
    driver = get_driver()
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = wb.create_sheet("house", 0)
    # 先填入第一列的欄位名稱
    sheet['A1'] = 'time'
    sheet['B1'] = 'text'
    sheet['C1'] = 'title'
    sheet['D1'] = 'area'
    sheet['E1'] = 'subway'
    sheet['F1'] = 'href'
    sheet['G1'] = 'style'

    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CLASS_NAME, 'item-title')))
    soup = bs4.BeautifulSoup(driver.page_source, 'html.parser')
    house = soup.find_all('a')
    # print(house)
    # print(len(house))

    for i in house:
        # print(i)
        data = i.find('div', class_='rent-item-right')
        # print(data)

        msg = i.find('div', class_='item-msg')
        price = i.find('div', class_='item-price-text')
        title = i.find('div', class_='item-title')
        area = i.find('div', class_='item-area')
        subway = i.find('div', class_='item-tip subway')
        link = ws.cell(row=1, column=1).value = '=HYPERLINK("{}")'.format(
            i.get('href'))
        style = i.find('ul', class_='item-style')

        if not title:
            pass
        else:
            if not ('林森北' in area.text):
                if not ('雅房' in title.text.strip()):
                    if not ('雅房' in style.text):
                        if not ('頂樓加蓋' in style.text):
                            if subway:
                                sheet.append([msg.text.strip()[49:55].replace('昨日', '1天前'), price.text, title.text.strip(), area.text, subway.text.strip(), link, style.text])
            wb.save("house.xlsx")


search()
